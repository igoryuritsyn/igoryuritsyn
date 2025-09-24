from telegram import Update, ReplyKeyboardMarkup, InputFile
from telegram.ext import (
    Application, CommandHandler, MessageHandler, ConversationHandler,
    ContextTypes, filters
)
from openpyxl import Workbook, load_workbook
import os
import mimetypes
from datetime import datetime

# --- НАСТРОЙКИ ---
TOKEN = "8396747361:AAFJKCy4kEXrVqnENDbpyqWVo8DtI9FksGQ"  # 🔴 замените на токен вашего бота
ADMIN_IDS = [1014020574, 1014020574]  # 🔴 список Telegram user_id админов
EXCEL_FILE = "bookings.xlsx"

# Этапы диалога
CHECKIN, CHECKOUT, ROOM_TYPE, CONFIRM = range(4)

# --- Работа с Excel ---
def save_booking(user_id, checkin, checkout, room_type):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["User ID", "Дата заезда", "Дата выезда", "Тип номера"])

    ws.append([user_id, checkin, checkout, room_type])
    wb.save(EXCEL_FILE)

def get_user_bookings(user_id):
    if not os.path.exists(EXCEL_FILE):
        return []

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    bookings = []
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == user_id:
            bookings.append({
                "row_index": row_index,
                "checkin": row[1],
                "checkout": row[2],
                "room_type": row[3]
            })
    return bookings

def delete_booking(row_index):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.delete_rows(row_index, 1)
    wb.save(EXCEL_FILE)

# --- Проверка дат ---
def is_valid_date(date_text):
    try:
        return datetime.strptime(date_text, "%Y-%m-%d")
    except ValueError:
        return None

# --- Команды ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Привет! Я бот для бронирования номеров в отеле 🏨\n\n"
        "Доступные команды:\n"
        "/book – забронировать номер\n"
        "/mybookings – мои брони\n"
        "/cancelbooking – отменить бронь\n"
        "/cancel – отменить процесс бронирования"
        + ("\n/admin: /allbookings – все брони (для администратора)" if update.effective_user.id in ADMIN_IDS else "")
    )

# Бронирование
async def book(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Введите дату заезда (например, 2025-09-25):")
    return CHECKIN

async def checkin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    checkin_date = is_valid_date(update.message.text)
    if not checkin_date:
        await update.message.reply_text("❌ Неверный формат даты. Используйте ГГГГ-ММ-ДД.")
        return CHECKIN

    today = datetime.today().date()
    if checkin_date.date() < today:
        await update.message.reply_text("❌ Дата заезда не может быть в прошлом. Введите другую дату.")
        return CHECKIN

    context.user_data["checkin"] = update.message.text
    await update.message.reply_text("Введите дату выезда (например, 2025-09-28):")
    return CHECKOUT

async def checkout(update: Update, context: ContextTypes.DEFAULT_TYPE):
    checkout_date = is_valid_date(update.message.text)
    if not checkout_date:
        await update.message.reply_text("❌ Неверный формат даты. Используйте ГГГГ-ММ-ДД.")
        return CHECKOUT

    checkin_date = is_valid_date(context.user_data["checkin"])
    if checkout_date <= checkin_date:
        await update.message.reply_text("❌ Дата выезда должна быть позже даты заезда. Введите снова.")
        return CHECKOUT

    context.user_data["checkout"] = update.message.text
    reply_keyboard = [["Одноместный", "Двухместный", "Люкс"]]
    await update.message.reply_text(
        "Выберите тип номера:",
        reply_markup=ReplyKeyboardMarkup(reply_keyboard, one_time_keyboard=True)
    )
    return ROOM_TYPE

async def room_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["room_type"] = update.message.text
    await update.message.reply_text(
        f"Подтвердите бронирование:\n"
        f"Дата заезда: {context.user_data['checkin']}\n"
        f"Дата выезда: {context.user_data['checkout']}\n"
        f"Тип номера: {context.user_data['room_type']}\n\n"
        f"Напишите 'Да' для подтверждения или 'Нет' для отмены."
    )
    return CONFIRM

async def confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text.lower() == "да":
        save_booking(
            user_id=update.effective_user.id,
            checkin=context.user_data["checkin"],
            checkout=context.user_data["checkout"],
            room_type=context.user_data["room_type"]
        )

        # Сообщение пользователю
        await update.message.reply_text("✅ Ваша бронь сохранена и подтверждена! Спасибо 🏨")

        # 🔔 Уведомления всем администраторам
        for admin_id in ADMIN_IDS:
            await context.bot.send_message(
                chat_id=admin_id,
                text=(
                    f"📢 Новая бронь!\n\n"
                    f"👤 Пользователь: {update.effective_user.full_name} (ID: {update.effective_user.id})\n"
                    f"📅 Заезд: {context.user_data['checkin']}\n"
                    f"📅 Выезд: {context.user_data['checkout']}\n"
                    f"🏨 Тип номера: {context.user_data['room_type']}"
                )
            )
    else:
        await update.message.reply_text("❌ Бронирование отменено.")
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Процесс бронирования отменён.")
    return ConversationHandler.END

# Мои брони
async def mybookings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    bookings = get_user_bookings(user_id)

    if not bookings:
        await update.message.reply_text("У вас пока нет броней ❌")
        return

    text = "Ваши бронирования:\n\n"
    for i, b in enumerate(bookings, start=1):
        text += (
            f"📌 {i}.\n"
            f"Дата заезда: {b['checkin']}\n"
            f"Дата выезда: {b['checkout']}\n"
            f"Тип номера: {b['room_type']}\n\n"
        )
    text += "Чтобы отменить бронь, используйте команду /cancelbooking"
    await update.message.reply_text(text)

# Отмена брони
async def cancelbooking(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    bookings = get_user_bookings(user_id)

    if not bookings:
        await update.message.reply_text("У вас нет активных броней ❌")
        return

    text = "Введите номер брони, которую хотите отменить:\n\n"
    for i, b in enumerate(bookings, start=1):
        text += (
            f"{i}. Заезд {b['checkin']} – Выезд {b['checkout']} ({b['room_type']})\n"
        )

    context.user_data["cancel_list"] = bookings
    await update.message.reply_text(text)

async def handle_cancel_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if "cancel_list" not in context.user_data:
        return

    try:
        choice = int(update.message.text.strip())
        bookings = context.user_data["cancel_list"]

        if 1 <= choice <= len(bookings):
            row_index = bookings[choice - 1]["row_index"]
            delete_booking(row_index)
            await update.message.reply_text("✅ Бронь успешно отменена!")
            context.user_data.pop("cancel_list", None)
        else:
            await update.message.reply_text("Неверный номер. Попробуйте снова.")
    except ValueError:
        await update.message.reply_text("Введите число (номер брони).")

# Все брони (только для админов)
async def allbookings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ У вас нет доступа к этой команде.")
        return

    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text("❌ Файл с бронями пока не создан.")
        return

    # Определяем правильный MIME-тип
    mime_type, _ = mimetypes.guess_type(EXCEL_FILE)
    if not mime_type:
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    with open(EXCEL_FILE, "rb") as f:
        await update.message.reply_document(
            document=InputFile(f, filename="bookings.xlsx"),
            caption="📂 Все брони"
        )

# --- MAIN ---
def main():
    app = Application.builder().token(TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("book", book)],
        states={
            CHECKIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, checkin)],
            CHECKOUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, checkout)],
            ROOM_TYPE: [MessageHandler(filters.TEXT & ~filters.COMMAND, room_type)],
            CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, confirm)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("mybookings", mybookings))
    app.add_handler(CommandHandler("cancelbooking", cancelbooking))
    app.add_handler(CommandHandler("allbookings", allbookings))
    app.add_handler(conv_handler)

    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_cancel_choice))

    print("Бот запущен...")
    app.run_polling()

if __name__ == "__main__":
    main()

